import openpyxl as excel


def write_in_cell(sheet, column, row, content):
    pass


data_set = [["abc", 1, 2, "abc12"],
            ["edf", 3, 2, "def32"],
            ["abc", 1, 3, "abc13"],
            ["ghi", 3, 4, "ghi34"]
            ]

book = excel.Workbook()

for data in data_set:
    # 配列の内容を変数に格納
    write_sheet = data[0]
    write_column = data[1]
    write_row = data[2]
    write_content = data[3]

    # excelシートに指定のシート名が存在するかをチェック
    if not write_sheet in book.sheetnames:
        book.create_sheet(write_sheet)

    # 書き込み
    sheet = book[write_sheet]
    cell = sheet.cell(row=write_row, column=write_column)
    cell.value = write_content

book.save("test_write_in_cell.xlsx")

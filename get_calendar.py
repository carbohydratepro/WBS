import calendar
import pstats
from traceback import print_stack
import jpholiday
import datetime
import openpyxl as excel


def main():
    def create_sheet():
        pass

    def write_in_cell():
        pass

    def data_process():
        pass

    def isBizDay(date):
        date = datetime.date(int(date[0:4]), int(date[4:6]), int(date[6:8]))
        if date.weekday() >= 5 or jpholiday.is_holiday(date):
            return 0
        else:
            return 1

    # 取得するカレンダーが日曜日から始まるよう設定
    calendar.setfirstweekday(calendar.SUNDAY)

    # 取得する年月をリストで格納
    get_month = [[2021, 12], [2022, 1], [2022, 2]]

    # 取得したカレンダーを格納するリストを宣言
    some_months = []

    # 取得した月の週数を格納
    week_month = []

    # 働く日のデータ格納
    work_date = []

    # 取得する月の数for文回す
    for i, gm in enumerate(get_month):  # 稼働日のデータを作成
        # カレンダーを取得
        month = calendar.monthcalendar(gm[0], gm[1])
        for week in month:
            for j, day in enumerate(week):
                if week[j] != 0:

                    s_year = str(gm[0])
                    if gm[1] < 10:
                        s_month = "0" + str(gm[1])
                    else:
                        s_month = str(gm[1])
                    if week[j] < 10:
                        s_day = "0" + str(week[j])
                    else:
                        s_day = str(week[j])

                    work_date.append([s_year + "/" + s_month + "/" + s_day])

    # 稼働日データの作成
    for i, date in enumerate(work_date):
        work_date[i].append(isBizDay(''.join( x for x in date[0] if x not in "/")))


    # 取得する月の数for文回す
    for i, gm in enumerate(get_month):
        # カレンダーを取得
        month = calendar.monthcalendar(gm[0], gm[1])

        # 取得したカレンダーの最後の日を格納
        end_day = max(month[-1])

        # 取得したカレンダーの配列の添え字の最大値を６にそろえる
        while (len(month) < 6):
            month.append([0]*7)

        # 取得した月の週の数を格納
        for _ in range(len(month)):
            week_month.append(get_month[i])

        # 取得したカレンダーをリストに格納
        for week in month:
            some_months.append(week)

    # 取得した年月日の祝日を取得
    holidays = jpholiday.between(
        datetime.date(get_month[0][0], get_month[0][1], 1), datetime.date(get_month[-1][0], get_month[-1][1], end_day))

    # 祝日をカレンダーに反映
    for i, (sm, wm) in enumerate(zip(some_months, week_month)):
        for j, d in enumerate(sm):
            for h in holidays:
                #some_months[i][j] : excelに書き込む元データ
                # wm : some_months[i][j]に対応する年月
                # h : [年月日（辞書型）, 祝日名]

                # some_months[i][j]が祝日だった場合
                if (wm[0] == h[0].year and wm[1] == h[0].month and d == h[0].day):
                    # some_months[i][j]に祝日名を追加する
                    some_months[i][j] = str(some_months[i][j]) + "\n" + h[1]

    # エクセルファイルを読み込み
    book = excel.load_workbook("カレンダー.xlsx")

    # calendarという名前のシートを選択
    sheet = book["data"]

    # sheet:dataに、取得したカレンダーを書き込み
    for i, week in enumerate(some_months):
        for j, day in enumerate(week):
            # 取得したカレンダーの存在しない部分を置き換え
            if day == 0:
                day = " "
            # シートの３行３列から順に書き込み
            cell = sheet.cell(row=3+i, column=3+j)
            cell.value = day

        # 書き込みしたカレンダーの情報を付随
        cell = sheet.cell(row=3+i, column=2)
        cell.value = str(week_month[i][0])+"/" + \
            str(week_month[i][1])+"_"+str(i % 6+1)

    # sheet:dataに、稼働日カレンダーを書き込み
    for i, date in enumerate(work_date):
        # シートの３行１５列から順に書き込み
        cell = sheet.cell(row=3+i, column=15)
        cell.value = date[0]
        cell = sheet.cell(row=3+i, column=16)
        cell.value = date[1]

    # カレンダーの取得年月日を書き込み
    for i, month in enumerate(get_month):
        cell = sheet.cell(row=3+i, column=10)
        cell.value = str(month[0]) + "/" + str(month[1])

    # 上書き保存
    book.save("カレンダー.xlsx")


if __name__ == '__main__':
    main()

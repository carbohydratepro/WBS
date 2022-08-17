import openpyxl as excel
import csv

book = excel.load_workbook('./test_write_csv.xlsx')

data_set = []

for s_name in book.sheetnames:
  sheet = book[s_name]
  for i in range(150): #row
    for j in range(50): #column
      content = sheet.cell(row=i+1, column=j+1).value

      #データの入力されているセルの情報のみを配列に格納
      if not content == None:
        data_set.append([s_name, j+1, i+1, content]) #シート名、列、行、内容


with open('data.csv', 'w') as file:
  writer = csv.writer(file,lineterminator='\n')
  writer.writerows(data_set)
